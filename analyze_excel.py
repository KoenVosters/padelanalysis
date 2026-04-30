import openpyxl
from openpyxl import load_workbook
import json

# Load the Excel file
wb = load_workbook('Padelscore v3.xlsx', data_only=False)

# Get basic info
print("=" * 60)
print("EXCEL FILE ANALYSIS - Padelscore v3.xlsx")
print("=" * 60)

print(f"\nSheet names: {wb.sheetnames}\n")

# Analyze each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*60}")
    print(f"SHEET: {sheet_name}")
    print(f"{'='*60}")
    print(f"Dimensions: {ws.dimensions}")
    
    # Display content
    print("\nContent:")
    for row in ws.iter_rows(min_row=1, max_row=min(30, ws.max_row), values_only=False):
        for cell in row:
            if cell.value is not None:
                value = cell.value
                # Check if it's a formula
                if isinstance(value, str) and value.startswith('='):
                    print(f"  {cell.coordinate}: FORMULA: {value}")
                else:
                    print(f"  {cell.coordinate}: {value}")
            else:
                print(f"  {cell.coordinate}: (empty)")

print("\n" + "="*60)
print("Analysis complete")
